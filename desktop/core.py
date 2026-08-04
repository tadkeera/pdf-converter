# -*- coding: utf-8 -*-
"""
PDF → Excel Conversion Engine
================================
Core engine used by both the desktop GUI (app.py) and the CLI.

Features
--------
* Converts ANY number of PDF pages (configurable limit, default up to 1000)
* Detects tables on each page (PyMuPDF find_tables) and exports them cell-per-cell
* Falls back to line-by-line text export when no table structure exists
* Extracts the merchant / agent name and uses it as the Excel file title
* Skips pages by page number ("1,3,5-7") and/or by page title / header keywords

Author: tadkeera (PDF Converter project)
"""

import os
import re
import math
import unicodedata
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from typing import List, Optional, Set, Tuple

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Arabic / bidi text normalisation
# ----------------------------------------------------------------------------
# PyMuPDF extracts RTL text in *visual* (display) order with presentation-form
# glyphs (ﻡ etc.). We normalise presentation forms back to base letters and
# restore the *logical* order so detection and Excel output behave correctly.

_ARABIC_RE = re.compile(r"[\u0600-\u06FF]")
_COMMON_AR_WORDS = (
    "اسم", "التاجر", "الوكيل", "المورد", "المستورد", "شركة", "مؤسسة", "فاتورة",
    "رقم", "تاريخ", "الكمية", "الإجمالي", "السعر", "الصنف", "الشهر", "القيمة",
    "عدد", "بضائع", "العميل", "التجارة", "العامة", "الدفع", "نقدا", "شروط",
    "الملخص", "ملخص", "جدول", "يناير", "فبراير", "مارس", "أبريل", "مايو",
    "يونيو", "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر",
    "البيان", "البيانات", "المجموع", "الصافي", "الضريبة", "الخصم",
)


def contains_arabic(text: str) -> bool:
    return bool(_ARABIC_RE.search(text or ""))


_NUMBER_RUN = re.compile(r"(?<![A-Za-z])[\d][\d.,:/\-]*(?![A-Za-z])")


def _fix_number_runs(text: str) -> str:
    """After a whole-line reversal, embedded numbers get mirrored
    (e.g. '2024-001' → '100-4202'). Reverse such digit runs back."""
    return _NUMBER_RUN.sub(lambda m: m.group()[::-1], text)


def to_logical(text: Optional[str]) -> str:
    """
    Convert a possibly visual-order Arabic string back to logical order:
    NFKC (presentation forms → base letters) + reverse when the line is
    clearly rendered RTL. Digit runs are repaired so numbers stay readable.
    """
    if not text:
        return ""
    s = unicodedata.normalize("NFKC", text).replace("\xa0", " ")
    if not contains_arabic(s):
        return s.strip()
    rev = s[::-1]

    def score(t: str) -> int:
        return sum(1 for w in _COMMON_AR_WORDS if w in t)

    # a label line rendered RTL almost always starts with ':' in visual order
    if s.lstrip().startswith(":") or score(rev) > score(s):
        return _fix_number_runs(rev).strip()
    return s.strip()

# ----------------------------------------------------------------------------
# Data structures
# ----------------------------------------------------------------------------

@dataclass
class PageData:
    """One processed PDF page."""
    number: int                       # original 1-based page number
    title: str                        # detected page title (used as sheet name)
    table: Optional[List[List[Optional[str]]]] = None   # extracted table (rows x cols)
    lines: Optional[List[str]] = None                   # fallback plain text lines


@dataclass
class ConversionResult:
    pages_converted: int = 0
    pages_skipped: int = 0
    pages_total: int = 0
    sheets: List[str] = field(default_factory=list)


# ----------------------------------------------------------------------------
# Small text helpers
# ----------------------------------------------------------------------------

_ILLEGAL_XML = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def clean_cell(value) -> str:
    """Normalise a cell value: string, strip control chars, cap Excel's 32767 limit."""
    if value is None:
        return ""
    text = to_logical(str(value))
    text = _ILLEGAL_XML.sub("", text)
    return text[:32767]


def parse_skip_numbers(spec: str) -> Set[int]:
    """Parse '1,3,5-9,12' into a set of page numbers."""
    result = set()
    if not spec:
        return result
    for part in str(spec).split(","):
        part = part.strip()
        if not part:
            continue
        m = re.fullmatch(r"(\d+)\s*-\s*(\d+)", part)
        if m:
            start, end = int(m.group(1)), int(m.group(2))
            if start <= end:
                result.update(range(start, end + 1))
        elif part.isdigit():
            result.add(int(part))
    return result


def parse_skip_keywords(spec: str) -> List[str]:
    """Parse comma-separated keywords."""
    return [k.strip() for k in str(spec).split(",") if k.strip()]


def sanitize_sheet_name(name: str, fallback: str = "ورقة") -> str:
    """Excel sheet names: <=31 chars and free of  : \\ / ? * [ ]  characters."""
    name = re.sub(r"[\\/?*\[\]:]", "_", name).strip()
    name = re.sub(r"\s+", " ", name)
    if not name:
        name = fallback
    return name[:31] or fallback


def sanitize_filename(name: Optional[str], fallback: str = "output") -> str:
    """Windows/Excel safe file name."""
    if not name:
        return fallback
    name = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", name).strip().strip(".")
    name = re.sub(r"\s+", " ", name).strip()
    if not name or name.lower() in {"con", "prn", "aux", "nul"}:
        name = fallback
    return name[:90]


def looks_like_label(text: str) -> bool:
    """Heuristic: is this text just a field label (no real name value)?"""
    t = text.strip(" :.-–—")
    if not t:
        return True
    return len(t) <= 2


# ----------------------------------------------------------------------------
# Page title detection
# ----------------------------------------------------------------------------

def _table_ok(table, lines: List[str], strategy: str) -> bool:
    """
    Validate a candidate table.
    - 'lines' strategy finds real ruled tables: trust them.
    - 'text' strategy can split Arabic words at whitespace gaps; only accept
      when the reconstructed rows match the page's actual text lines.
    """
    if strategy == "lines":
        return len(table) >= 1
    if len(table) < 2 or len(table[0]) < 2:
        return False
    if not lines:
        return True
    line_norms = [re.sub(r"\s+", " ", ln) for ln in lines if ln.strip()]
    hits = checked = 0
    for row in table:
        joined = re.sub(
            r"\s+", " ",
            " ".join(str(c) for c in row if c is not None and str(c).strip()),
        ).strip()
        if not joined:
            continue
        checked += 1
        if any(SequenceMatcher(None, joined, ln).ratio() >= 0.85 for ln in line_norms):
            hits += 1
    return hits / max(checked, 1) >= 0.7


def detect_page_title(lines: List[str], page_no: int, max_len: int = 60) -> str:
    """The page title is its first short meaningful line; else 'Page N'."""
    for ln in lines:
        ln = ln.strip()
        if not ln or looks_like_label(ln):
            continue
        if len(ln) <= max_len:
            return ln
        break
    return f"Page {page_no}"


# ----------------------------------------------------------------------------
# Extraction
# ----------------------------------------------------------------------------

def extract_pages(
    pdf_path: str,
    max_pages: Optional[int] = None,
    skip_numbers: Optional[Set[int]] = None,
    skip_keywords: Optional[List[str]] = None,
    skip_blank: bool = True,
) -> Tuple[List[PageData], ConversionResult]:
    """Extract every page of a PDF into PageData objects."""
    skip_numbers = skip_numbers or set()
    skip_keywords = [k.lower() for k in (skip_keywords or []) if k.strip()]
    result = ConversionResult()

    doc = fitz.open(pdf_path)
    result.pages_total = doc.page_count
    total = doc.page_count if not max_pages else min(doc.page_count, max_pages)

    pages: List[PageData] = []
    for i in range(total):
        page_no = i + 1
        if page_no in skip_numbers:
            result.pages_skipped += 1
            continue

        page = doc.load_page(i)
        raw_text = page.get_text("text") or ""
        lines = [ln for ln in (to_logical(l) for l in raw_text.splitlines()) if ln]
        title = detect_page_title(lines, page_no)

        # Skip blank pages (no text, no images, no drawings)
        if skip_blank and not lines:
            try:
                has_images = len(page.get_images(full=True)) > 0
                has_drawings = len(page.get_drawings()) > 0
            except Exception:
                has_images = has_drawings = False
            if not has_images and not has_drawings:
                result.pages_skipped += 1
                continue

        # Skip by keyword in the page title / first lines of the page
        if skip_keywords:
            probe = " ".join(lines[:15]).lower()
            if any(k in probe for k in skip_keywords):
                result.pages_skipped += 1
                continue

        # 1) Try real table detection (ruled-line tables first, then text-only)
        table = None
        try:
            for strategy in ("lines", "text"):
                tabs = page.find_tables(strategy=strategy)
                cands = tabs.tables
                if not cands:
                    continue
                candidate = [row for row in cands[0].extract()]
                if not _table_ok(candidate, lines, strategy):
                    continue
                # keep the richest valid candidate (most rows) unless trivial
                if len(candidate) > (len(table) if table else 1):
                    table = candidate
                if len(candidate) >= 2:
                    break
        except Exception:
            table = None

        if table:
            pages.append(PageData(page_no, title, table=table))
        else:
            # 2) Fallback: text lines as single-column rows
            pages.append(PageData(page_no, title, lines=lines))

        result.pages_converted += 1

    doc.close()
    return pages, result


# ----------------------------------------------------------------------------
# Excel writer
# ----------------------------------------------------------------------------

_TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
_TITLE_FONT = Font(name="Arial", size=14, bold=True, color="FFFFFF")
_HEADER_FONT = Font(name="Arial", size=11, bold=True, color="1F4E79")
_CELL_FONT = Font(name="Arial", size=10)
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")
_WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def write_excel(
    pages: List[PageData],
    out_path: str,
    merchant: Optional[str] = None,
    source_pdf: Optional[str] = None,
) -> List[str]:
    """Write the pages into one .xlsx workbook. Returns the sheet names."""
    wb = Workbook()
    first_sheet = True
    sheet_names: List[str] = []

    for pdata in pages:
        ws = wb.active if first_sheet else wb.create_sheet()
        first_sheet = False
        ws.title = sanitize_sheet_name(pdata.title, fallback=f"Page {pdata.number}")
        sheet_names.append(ws.title)
        ws.sheet_view.rightToLeft = True  # proper Arabic / RTL support
        ws.sheet_view.showGridLines = False

        # ---- Title band (row 1) ------------------------------------------
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
        title_cell = ws.cell(row=1, column=1)
        parts = []
        if merchant:
            parts.append(merchant)
        parts.append(f"{pdata.title}  (الصفحة {pdata.number})")
        title_cell.value = "  |  ".join(parts)
        title_cell.font = _TITLE_FONT
        title_cell.fill = _TITLE_FILL
        title_cell.alignment = _WRAP_CENTER
        ws.row_dimensions[1].height = 30

        # ---- Data ----------------------------------------------------------
        start_row = 2
        if pdata.table:
            ncols = max((len(r) for r in pdata.table), default=1)
            for r_idx, row in enumerate(pdata.table):
                for c_idx in range(ncols):
                    value = clean_cell(row[c_idx]) if c_idx < len(row) else ""
                    cell = ws.cell(row=start_row + r_idx, column=c_idx + 1, value=value)
                    cell.font = _HEADER_FONT if r_idx == 0 else _CELL_FONT
                    if r_idx == 0:
                        cell.fill = _HEADER_FILL
                    cell.border = _BORDER
                    cell.alignment = _WRAP
            # auto-width (capped)
            for c_idx in range(ncols):
                letter = get_column_letter(c_idx + 1)
                best = max(
                    (len(clean_cell(row[c_idx])) for row in pdata.table if c_idx < len(row)),
                    default=8,
                )
                ws.column_dimensions[letter].width = min(max(best * 1.35, 9), 60)
        elif pdata.lines:
            for r_idx, line in enumerate(pdata.lines):
                cell = ws.cell(row=start_row + r_idx, column=1, value=clean_cell(line))
                cell.font = _CELL_FONT
                cell.alignment = _WRAP
                cell.border = _BORDER
            ws.column_dimensions["A"].width = 110

    # Workbook properties
    wb.properties.creator = "PDF Converter"
    if merchant:
        wb.properties.title = merchant
    if source_pdf:
        wb.properties.subject = f"Source: {os.path.basename(source_pdf)}"

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return sheet_names


# ----------------------------------------------------------------------------
# High-level entry point (used by GUI and CLI)
# ----------------------------------------------------------------------------

def convert_pdf(
    pdf_path: str,
    out_dir: str,
    output_filename: Optional[str] = None,
    merchant: Optional[str] = None,
    max_pages: Optional[int] = 1000,
    skip_numbers: Optional[Set[int]] = None,
    skip_keywords: Optional[List[str]] = None,
    skip_blank: bool = True,
    progress_cb=None,
) -> Tuple[str, ConversionResult]:
    """Convert one PDF file into one .xlsx file."""
    pages, result = extract_pages(
        pdf_path,
        max_pages=max_pages,
        skip_numbers=skip_numbers,
        skip_keywords=skip_keywords,
        skip_blank=skip_blank,
    )
    if progress_cb:
        progress_cb(0.5)

    if not pages:
        raise ValueError("لا توجد صفحات للتحويل (ربما تم استثناء كل الصفحات)")

    if not output_filename:
        base = sanitize_filename(os.path.splitext(os.path.basename(pdf_path))[0])
        output_filename = f"{base}.xlsx"
    if not output_filename.lower().endswith(".xlsx"):
        output_filename += ".xlsx"
    # Never allow path separators to leak into the file name
    output_filename = sanitize_filename(os.path.basename(output_filename))

    out_path = os.path.join(out_dir, output_filename)
    sheet_names = write_excel(pages, out_path, merchant=merchant, source_pdf=pdf_path)
    result.sheets = sheet_names
    if progress_cb:
        progress_cb(1.0)
    return out_path, result


# ----------------------------------------------------------------------------
# CLI (for headless use / testing)
# ----------------------------------------------------------------------------

def main(argv=None):
    import argparse
    import sys

    p = argparse.ArgumentParser(
        description="PDF → Excel converter (tadkeera). Converts every PDF page "
                    "into a sheet, extracts the merchant/agent name as file title."
    )
    p.add_argument("pdf", nargs="+", help="PDF file(s) to convert")
    p.add_argument("-o", "--outdir", default=".", help="Output directory")
    p.add_argument("--max-pages", type=int, default=1000, help="Max pages per file (default 1000)")
    p.add_argument("--skip", default="", help="Pages to skip: 1,3,5-9")
    p.add_argument("--skip-kw", default="", help="Keywords: skip pages whose title/header contains them")
    p.add_argument("--no-skip-blank", action="store_true", help="Keep blank pages")
    p.add_argument("--name", default="", help="Manual merchant/agent name (overrides detection)")
    args = p.parse_args(argv)

    from extractor import extract_merchant_name

    skip_nums = parse_skip_numbers(args.skip)
    skip_kws = parse_skip_keywords(args.skip_kw)

    for pdf in args.pdf:
        if not os.path.exists(pdf):
            print(f"✗ {pdf}: file not found", file=sys.stderr)
            continue
        merchant = args.name or extract_merchant_name(pdf)
        fname = sanitize_filename(merchant) if merchant else None
        try:
            out, res = convert_pdf(
                pdf,
                args.outdir,
                output_filename=fname,
                merchant=merchant,
                max_pages=args.max_pages,
                skip_numbers=skip_nums,
                skip_keywords=skip_kws,
                skip_blank=not args.no_skip_blank,
            )
            print(f"✓ {os.path.basename(pdf)} → {out}  "
                  f"({res.pages_converted} pages, {res.pages_skipped} skipped)")
            if merchant:
                print(f"  التاجر/الوكيل المستخرج: {merchant}")
        except Exception as e:
            print(f"✗ {pdf}: {e}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
